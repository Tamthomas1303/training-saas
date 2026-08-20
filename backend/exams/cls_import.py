"""
Import ngan hang cau hoi tu file Excel xuat tu CLS (Prompt_Import_CauHoi_CLS.md) - doc sheet
"Cau Hoi", tao QuestionBank (gom theo Chu De) + Question + QuestionOption. Toan bo ngan hang CLS
(~2.130 cau, da xac nhan) CHI gom trac nghiem 1 dap an + nhieu dap an (khong co Gach chan/Menh
de/Cau chum) - dang khac gap thi LOG CANH BAO + BO QUA, khong tao (phong ho, khong ky vong xay ra).

Cau truc file thuc te (kiem tra truc tiep tren "Ngan hang cau hoi phuc vu Kampong.xlsx"): dong 1
la tieu de gop o ("MAU NGAN HANG CAU HOI"), dong 2 moi la hang tieu de cot THAT (co nhieu khoang
trang thua quanh moi ten cot - vd '      STT (*)      ' - PHAI .strip()), du lieu tu dong 3. De
AN TOAN voi cac file chu de khac (Pha che/Thu ngan/Coca...) co the lech 1 dong, ham
_find_header_row() TU DO tim dong tieu de that (khong gia dinh cung dong 2) bang cach do 1 dong
co chua ca 2 cot bat buoc 'Noi Dung' va 'Dap An Dung' (sau khi da chuan hoa - bo khoang trang +
hau to ' (*)').

Dap An Dung: la SO THU TU (1-based) cua cot "Cau Tra Loi N" - single la 1 so, multiple la danh
sach so ngan cach dau phay (vd "1, 3"). Index tinh THEO VI TRI COT GOC (1..8), khong phai vi tri
sau khi da loc bo cac o trong - dung y prompt "chu y index 1-based" va khop voi cach nguoi nhap
lieu CLS danh so cot tren form.

Cot "Nang luc" (TUY CHON, Prompt_GanNangLuc_CauHoi_Excel.md muc 3): neu file co them cot nay (anh
tu them tay vao file CLS xuat ra) thi gan luon competency khi import, khop TEN da chuan hoa
(dung chung logic voi exams/competency_assign.py::apply_competency_assignments). Khong co cot ->
bo qua nhu truoc (competency=None, gan sau qua man Xuat/Nhap Excel gan nang luc).
"""
import re

from .models import Question, QuestionBank, QuestionOption

SHEET_NAME = 'Câu Hỏi'

# Cac ten cot BAT BUOC de nhan dien dung dong tieu de that (sau khi chuan hoa - xem _clean_header).
REQUIRED_HEADERS = {'Nội Dung', 'Đáp Án Đúng', 'Cấp Độ', 'Kiểu Câu Hỏi', 'Chủ Đề'}

TYPE_MAP = {
    'Trắc nghiệm một lựa chọn': Question.Type.SINGLE,
    'Trắc nghiệm nhiều lựa chọn': Question.Type.MULTIPLE,
}

# 4 cap do CLS -> 3 muc Question.Difficulty. "Thong Hieu" xep vao medium (cung Van Dung) thay vi
# easy - hop ly hon ve mat ngu nghia (thong hieu > nhan biet), co the dieu chinh sau neu can.
LEVEL_MAP = {
    'Nhận biết': Question.Difficulty.EASY,
    'Thông Hiểu': Question.Difficulty.MEDIUM,
    'Vận Dụng': Question.Difficulty.MEDIUM,
    'Vận Dụng Cao': Question.Difficulty.HARD,
}

ANSWER_COLUMNS = [f'Câu Trả Lời {i}' for i in range(1, 9)]


def _clean_header(value):
    text = (value or '').strip() if isinstance(value, str) else ''
    return re.sub(r'\s*\(\*\)\s*$', '', text)


def _clean_cell(value):
    if value is None:
        return ''
    return str(value).strip()


def _find_header_row(rows, max_scan=5):
    """Tra ve (index trong `rows`, {ten_cot_da_chuan_hoa: vi_tri_cot}) cua dong tieu de that -
    dong DAU TIEN (trong may dong dau) co du cac REQUIRED_HEADERS sau khi chuan hoa. None neu
    khong tim thay (file sai dinh dang)."""
    for i, row in enumerate(rows[:max_scan]):
        header = {_clean_header(c): j for j, c in enumerate(row) if _clean_header(c)}
        if REQUIRED_HEADERS.issubset(header.keys()):
            return i, header
    return None, None


def _parse_correct_indices(raw):
    """'2' -> {2}; '1, 3' -> {1, 3}; rong/khong parse duoc -> set() (row se bi bo qua)."""
    indices = set()
    for token in (raw or '').split(','):
        token = token.strip()
        if not token:
            continue
        try:
            indices.add(int(token))
        except ValueError:
            return set()
    return indices


def parse_workbook(path_or_file, sheet_name=SHEET_NAME):
    """Doc file Excel CLS -> {'bank_name': [rows...]}  cac row da parse (dict) + 'skipped':
    [{'row': stt_hoac_vi_tri, 'reason': ...}, ...]. KHONG dong DB - thuan tuy doc + validate, de
    dung chung cho ca --dry-run lan import that (import_rows ben duoi moi ghi DB)."""
    import openpyxl

    wb = openpyxl.load_workbook(path_or_file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong file (có: {', '.join(wb.sheetnames)}).")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_idx, col = _find_header_row(rows)
    if header_idx is None:
        raise ValueError(
            "Không tìm được dòng tiêu đề (cần đủ cột STT/Chủ Đề/Kiểu Câu Hỏi/Nội Dung/Đáp Án Đúng/Cấp Độ)."
        )

    def get(row, name, default=''):
        i = col.get(name)
        if i is None or i >= len(row):
            return default
        return row[i]

    parsed = []
    skipped = []
    for row in rows[header_idx + 1:]:
        if row is None or not any(c is not None and str(c).strip() for c in row):
            continue  # dong trong hoan toan - bo qua am tham, khong tinh la "loi"

        stt = _clean_cell(get(row, 'STT')) or '?'
        content = _clean_cell(get(row, 'Nội Dung'))
        if not content:
            skipped.append({'row': stt, 'reason': 'Thiếu Nội Dung'})
            continue

        type_raw = _clean_cell(get(row, 'Kiểu Câu Hỏi'))
        q_type = TYPE_MAP.get(type_raw)
        if not q_type:
            skipped.append({'row': stt, 'reason': f"Kiểu câu hỏi không hỗ trợ: '{type_raw}'"})
            continue

        correct_indices = _parse_correct_indices(_clean_cell(get(row, 'Đáp Án Đúng')))
        if not correct_indices:
            skipped.append({'row': stt, 'reason': 'Thiếu/không đọc được Đáp Án Đúng'})
            continue

        options = []
        for pos, col_name in enumerate(ANSWER_COLUMNS, start=1):
            text = _clean_cell(get(row, col_name))
            if not text:
                continue
            options.append({'text': text, 'is_correct': pos in correct_indices})

        if not options:
            skipped.append({'row': stt, 'reason': 'Không có đáp án nào (Câu Trả Lời 1..8 trống)'})
            continue
        if not any(o['is_correct'] for o in options):
            skipped.append({'row': stt, 'reason': 'Đáp Án Đúng không khớp với vị trí đáp án nào có nội dung'})
            continue

        bank_name = _clean_cell(get(row, 'Chủ Đề')) or 'Chưa phân loại (CLS)'
        level_raw = _clean_cell(get(row, 'Cấp Độ'))
        difficulty = LEVEL_MAP.get(level_raw, Question.Difficulty.MEDIUM)

        parsed.append({
            'stt': stt, 'bank_name': bank_name, 'type': q_type,
            'stem_html': content, 'explanation_html': _clean_cell(get(row, 'Giải Thích Kết Quả')),
            'media_url': _clean_cell(get(row, 'Đường Dẫn Tệp Tin')),
            'difficulty': difficulty, 'options': options,
            'competency_name': _clean_cell(get(row, 'Năng lực')),  # rong neu file khong co cot nay
        })

    return {'parsed': parsed, 'skipped': skipped}


def import_rows(tenant, parsed_rows, dry_run=True):
    """Ghi (hoac chi dem, neu dry_run) QuestionBank/Question/QuestionOption tu ket qua
    parse_workbook()['parsed']. Idempotent: 1 Question da ton tai (cung tenant + bank + stem_html
    - dedup THEO NOI DUNG, dung y prompt "khong co truong luu STT nguon, khong can migration")
    thi BO QUA, khong tao lai/khong sua. Tra ve thong ke."""
    from .competency_assign import build_competency_index
    from dashboard.services import _normalize_competency_name

    stats = {
        'banks_created': 0, 'banks_existing': 0,
        'questions_created': 0, 'questions_skipped_duplicate': 0,
        'single_created': 0, 'multiple_created': 0, 'options_created': 0,
        'competency_matched': 0, 'competency_unmatched': 0,
    }
    bank_cache = {}
    competency_index = build_competency_index(tenant)

    for row in parsed_rows:
        bank_name = row['bank_name']
        if bank_name not in bank_cache:
            bank = QuestionBank.objects.filter(tenant=tenant, name=bank_name).first()
            if bank:
                stats['banks_existing'] += 1
            else:
                stats['banks_created'] += 1
                if not dry_run:
                    bank = QuestionBank.objects.create(tenant=tenant, name=bank_name)
            bank_cache[bank_name] = bank

        bank = bank_cache[bank_name]
        already_exists = bool(bank) and Question.objects.filter(
            tenant=tenant, bank=bank, stem_html=row['stem_html'],
        ).exists()
        if already_exists:
            stats['questions_skipped_duplicate'] += 1
            continue

        stats['questions_created'] += 1
        stats['options_created'] += len(row['options'])
        if row['type'] == Question.Type.SINGLE:
            stats['single_created'] += 1
        else:
            stats['multiple_created'] += 1

        competency = None
        competency_name = row.get('competency_name')
        if competency_name:
            matches = competency_index.get(_normalize_competency_name(competency_name)) or []
            if len(matches) == 1:
                competency = matches[0]
                stats['competency_matched'] += 1
            else:
                stats['competency_unmatched'] += 1

        if dry_run:
            continue

        question = Question.objects.create(
            tenant=tenant, bank=bank, type=row['type'], stem_html=row['stem_html'],
            explanation_html=row['explanation_html'], media_url=row['media_url'],
            difficulty=row['difficulty'], competency=competency,
        )
        QuestionOption.objects.bulk_create([
            QuestionOption(
                tenant=tenant, question=question, order=i,
                content_html=opt['text'], is_correct=opt['is_correct'],
            )
            for i, opt in enumerate(row['options'])
        ])

    return stats
