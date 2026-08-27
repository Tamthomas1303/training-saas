from datetime import timedelta
from decimal import Decimal
from unittest.mock import patch

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework.test import APIClient

from accounts.models import Tenant, User
from cls_sync.models import ExamResult
from dashboard.models import Competency, CompetencyGroup
from employees.models import Employee
from integration.models import XapiStatement
from restaurants.models import Restaurant

from .models import (
    Answer,
    Assessment,
    AssessmentAssignment,
    AssessmentQuestion,
    Attempt,
    ExamSession,
    ProctoringEvent,
    Question,
    QuestionBank,
    QuestionOption,
)
from .services import grade_objective_answer


class QuestionBankAdminApiTests(TestCase):
    """CRUD ngan hang cau hoi - chi Admin duoc ghi, moi role dang nhap doc duoc."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        self.client = APIClient()

    def test_admin_can_create_bank(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.post(reverse('exam-bank-list'), {'name': 'An toàn thực phẩm'})
        self.assertEqual(resp.status_code, 201)
        bank = QuestionBank.objects.get(id=resp.data['id'])
        self.assertEqual(bank.tenant_id, self.tenant.id)

    def test_non_admin_cannot_create_bank(self):
        self.client.force_authenticate(self.trainer)
        resp = self.client.post(reverse('exam-bank-list'), {'name': 'X'})
        self.assertEqual(resp.status_code, 403)

    def test_non_admin_can_read_banks(self):
        QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.client.force_authenticate(self.trainer)
        resp = self.client.get(reverse('exam-bank-list'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['count'], 1)


class QuestionAdminApiTests(TestCase):
    """Tao cau hoi single kem options long trong body (dung serializer sync options)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.client = APIClient()

    def test_create_single_question_with_options(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.post(reverse('exam-question-list'), {
            'bank': self.bank.id, 'type': Question.Type.SINGLE, 'stem_html': 'Thủ đô VN?', 'points': 2,
            'options': [
                {'content_html': 'Hà Nội', 'is_correct': True},
                {'content_html': 'Đà Nẵng', 'is_correct': False},
            ],
        }, format='json')
        self.assertEqual(resp.status_code, 201)
        question = Question.objects.get(id=resp.data['id'])
        self.assertEqual(question.options.count(), 2)
        self.assertEqual(question.options.filter(is_correct=True).count(), 1)


class ObjectiveGradingTests(TestCase):
    """Cham diem CA 8 DANG - test truc tiep services.grade_objective_answer, dac biet
    matching/dragdrop/multiple all-or-nothing, text_fill khong phan biet hoa/thuong, numeric
    trong nguong sai so."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')

    def _make_question(self, qtype, **kwargs):
        return Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=qtype, stem_html='Câu hỏi', **kwargs,
        )

    def test_single_correct_and_incorrect(self):
        q = self._make_question(Question.Type.SINGLE)
        correct = QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='A', is_correct=True)
        wrong = QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='B', is_correct=False)

        score, is_correct = grade_objective_answer(q, {'option_id': correct.id}, 2)
        self.assertEqual(score, Decimal('2'))
        self.assertTrue(is_correct)

        score, is_correct = grade_objective_answer(q, {'option_id': wrong.id}, 2)
        self.assertEqual(score, Decimal('0'))
        self.assertFalse(is_correct)

    def test_truefalse(self):
        q = self._make_question(Question.Type.TRUEFALSE)
        true_opt = QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='Đúng', is_correct=True)
        QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='Sai', is_correct=False)

        score, is_correct = grade_objective_answer(q, {'option_id': true_opt.id}, 1)
        self.assertEqual(score, Decimal('1'))
        self.assertTrue(is_correct)

    def test_multiple_all_or_nothing(self):
        q = self._make_question(Question.Type.MULTIPLE)
        a = QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='A', is_correct=True)
        b = QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='B', is_correct=True)
        c = QuestionOption.objects.create(tenant=self.tenant, question=q, content_html='C', is_correct=False)

        score, is_correct = grade_objective_answer(q, {'option_ids': [a.id, b.id]}, 3)
        self.assertEqual(score, Decimal('3'))
        self.assertTrue(is_correct)

        # thieu 1 dap an dung -> sai het (khong cham 1 phan)
        score, is_correct = grade_objective_answer(q, {'option_ids': [a.id]}, 3)
        self.assertEqual(score, Decimal('0'))
        self.assertFalse(is_correct)

        # dung nhung thua 1 dap an sai -> sai het
        score, is_correct = grade_objective_answer(q, {'option_ids': [a.id, b.id, c.id]}, 3)
        self.assertEqual(score, Decimal('0'))
        self.assertFalse(is_correct)

    def test_text_fill_case_insensitive(self):
        q = self._make_question(
            Question.Type.TEXT_FILL, config={'accepted': ['Paris'], 'case_sensitive': False},
        )
        score, is_correct = grade_objective_answer(q, {'text': 'paris'}, 1)
        self.assertEqual(score, Decimal('1'))
        self.assertTrue(is_correct)

        score, is_correct = grade_objective_answer(q, {'text': '  PARIS  '}, 1)
        self.assertTrue(is_correct)

        score, is_correct = grade_objective_answer(q, {'text': 'london'}, 1)
        self.assertFalse(is_correct)

    def test_text_fill_case_sensitive(self):
        q = self._make_question(
            Question.Type.TEXT_FILL, config={'accepted': ['Paris'], 'case_sensitive': True},
        )
        score, is_correct = grade_objective_answer(q, {'text': 'paris'}, 1)
        self.assertFalse(is_correct)

    def test_numeric_within_tolerance(self):
        q = self._make_question(Question.Type.NUMERIC, config={'answer': 10, 'tolerance': 0.5})
        score, is_correct = grade_objective_answer(q, {'value': 10.5}, 1)
        self.assertTrue(is_correct)
        score, is_correct = grade_objective_answer(q, {'value': 9.5}, 1)
        self.assertTrue(is_correct)

    def test_numeric_outside_tolerance(self):
        q = self._make_question(Question.Type.NUMERIC, config={'answer': 10, 'tolerance': 0.5})
        score, is_correct = grade_objective_answer(q, {'value': 10.6}, 1)
        self.assertFalse(is_correct)
        self.assertEqual(score, Decimal('0'))

    def test_matching_all_or_nothing(self):
        q = self._make_question(
            Question.Type.MATCHING,
            config={'pairs': [{'left': 'A', 'right': '1'}, {'left': 'B', 'right': '2'}]},
        )
        score, is_correct = grade_objective_answer(
            q, {'pairs': [{'left': 'A', 'right': '1'}, {'left': 'B', 'right': '2'}]}, 2,
        )
        self.assertTrue(is_correct)
        self.assertEqual(score, Decimal('2'))

        score, is_correct = grade_objective_answer(
            q, {'pairs': [{'left': 'A', 'right': '2'}, {'left': 'B', 'right': '1'}]}, 2,
        )
        self.assertFalse(is_correct)
        self.assertEqual(score, Decimal('0'))

    def test_dragdrop_all_or_nothing(self):
        q = self._make_question(
            Question.Type.DRAGDROP,
            config={'tokens': ['x', 'y'], 'gaps': [{'id': 1, 'answer': 'x'}, {'id': 2, 'answer': 'y'}]},
        )
        score, is_correct = grade_objective_answer(q, {'placements': {'1': 'x', '2': 'y'}}, 2)
        self.assertTrue(is_correct)

        score, is_correct = grade_objective_answer(q, {'placements': {'1': 'y', '2': 'x'}}, 2)
        self.assertFalse(is_correct)
        self.assertEqual(score, Decimal('0'))

    def test_essay_not_auto_graded(self):
        q = self._make_question(Question.Type.ESSAY)
        score, is_correct = grade_objective_answer(q, {'text': 'bài làm của tôi'}, 5)
        self.assertIsNone(score)
        self.assertIsNone(is_correct)


class ExamFlowApiTests(TestCase):
    """Smoke test theo muc 4 cua prompt: 1 ngan hang + 8 cau (moi dang 1 cau) -> 1 de 8 cau ->
    gan 1 nhan su -> lam bai -> tu cham 7 cau khach quan -> cham tay cau tu luan -> xem ket qua."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')

        self.q_single = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Q single',
        )
        self.opt_single_correct = QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_single, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_single, content_html='Sai', is_correct=False,
        )

        self.q_multiple = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.MULTIPLE, stem_html='Q multiple',
        )
        self.opt_m1 = QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_multiple, content_html='A', is_correct=True,
        )
        self.opt_m2 = QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_multiple, content_html='B', is_correct=True,
        )
        QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_multiple, content_html='C', is_correct=False,
        )

        self.q_truefalse = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.TRUEFALSE, stem_html='Q truefalse',
        )
        self.opt_tf_true = QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_truefalse, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(
            tenant=self.tenant, question=self.q_truefalse, content_html='Sai', is_correct=False,
        )

        self.q_text_fill = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.TEXT_FILL, stem_html='Q text_fill',
            config={'accepted': ['Paris'], 'case_sensitive': False},
        )
        self.q_numeric = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.NUMERIC, stem_html='Q numeric',
            config={'answer': 10, 'tolerance': 0.5},
        )
        self.q_matching = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.MATCHING, stem_html='Q matching',
            config={'pairs': [{'left': 'A', 'right': '1'}, {'left': 'B', 'right': '2'}]},
        )
        self.q_dragdrop = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.DRAGDROP, stem_html='Q dragdrop',
            config={'tokens': ['x', 'y'], 'gaps': [{'id': 1, 'answer': 'x'}, {'id': 2, 'answer': 'y'}]},
        )
        self.q_essay = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.ESSAY, stem_html='Q essay',
        )

        self.questions = [
            self.q_single, self.q_multiple, self.q_truefalse, self.q_text_fill, self.q_numeric,
            self.q_matching, self.q_dragdrop, self.q_essay,
        ]
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề kiểm tra tổng hợp', status=Assessment.Status.PUBLISHED,
            max_attempts=1, pass_mark=80, created_by=self.admin,
        )
        for i, q in enumerate(self.questions):
            AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=q, order=i)

        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)

        self.client = APIClient()

    def _correct_response(self, question):
        if question.id == self.q_single.id:
            return {'option_id': self.opt_single_correct.id}
        if question.id == self.q_multiple.id:
            return {'option_ids': [self.opt_m1.id, self.opt_m2.id]}
        if question.id == self.q_truefalse.id:
            return {'option_id': self.opt_tf_true.id}
        if question.id == self.q_text_fill.id:
            return {'text': 'paris'}
        if question.id == self.q_numeric.id:
            return {'value': 10.3}
        if question.id == self.q_matching.id:
            return {'pairs': [{'left': 'A', 'right': '1'}, {'left': 'B', 'right': '2'}]}
        if question.id == self.q_dragdrop.id:
            return {'placements': {'1': 'x', '2': 'y'}}
        return {'text': 'Bài làm tự luận của tôi.'}

    def test_full_flow_start_answer_submit_grade(self):
        self.client.force_authenticate(self.learner_user)

        start_resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(start_resp.status_code, 200)
        attempt_id = start_resp.data['attempt_id']
        self.assertEqual(len(start_resp.data['questions']), 8)
        # khong lo dap an dung ra ngoai
        for q_payload in start_resp.data['questions']:
            if q_payload['type'] in ('single', 'multiple', 'truefalse'):
                for opt in q_payload['options']:
                    self.assertNotIn('is_correct', opt)

        attempt = Attempt.objects.get(id=attempt_id)
        self.assertEqual(attempt.status, Attempt.Status.IN_PROGRESS)
        self.assertEqual(len(attempt.question_ids), 8)

        items = [{'question': q.id, 'response': self._correct_response(q)} for q in self.questions]
        save_resp = self.client.post(
            reverse('exam-attempt-answer', args=[attempt.id]), {'items': items}, format='json',
        )
        self.assertEqual(save_resp.status_code, 200)
        self.assertEqual(save_resp.data['saved'], 8)
        self.assertEqual(Answer.objects.filter(attempt=attempt).count(), 8)

        submit_resp = self.client.post(reverse('exam-attempt-submit', args=[attempt.id]))
        self.assertEqual(submit_resp.status_code, 200)
        self.assertEqual(submit_resp.data['status'], Attempt.Status.SUBMITTED)
        self.assertEqual(submit_resp.data['score'], Decimal('7.00'))
        self.assertEqual(submit_resp.data['max_score'], Decimal('8.00'))

        attempt.refresh_from_db()
        self.assertEqual(attempt.status, Attempt.Status.SUBMITTED)

        # nop lai lan 2 se bi tu choi
        resubmit_resp = self.client.post(reverse('exam-attempt-submit', args=[attempt.id]))
        self.assertEqual(resubmit_resp.status_code, 400)

        # danh sach cham tay hien attempt nay
        self.client.force_authenticate(self.admin)
        grading_resp = self.client.get(reverse('exam-grading'))
        self.assertEqual(len(grading_resp.data), 1)
        self.assertEqual(grading_resp.data[0]['id'], attempt.id)

        grade_resp = self.client.post(
            reverse('exam-attempt-grade', args=[attempt.id]),
            {'scores': {str(self.q_essay.id): 1}}, format='json',
        )
        self.assertEqual(grade_resp.status_code, 200)
        self.assertEqual(grade_resp.data['status'], Attempt.Status.GRADED)
        self.assertEqual(grade_resp.data['score'], Decimal('8.00'))
        self.assertEqual(grade_resp.data['percent'], Decimal('100.00'))
        self.assertTrue(grade_resp.data['passed'])

        assignment = AssessmentAssignment.objects.get(assessment=self.assessment, employee=self.employee)
        self.assertEqual(assignment.status, AssessmentAssignment.Status.DONE)

    def test_max_attempts_enforced(self):
        self.client.force_authenticate(self.learner_user)
        first = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(first.status_code, 200)
        # nop luon lan 1 de duoc phep bat dau lan tiep theo (khong con IN_PROGRESS)
        self.client.post(reverse('exam-attempt-submit', args=[first.data['attempt_id']]))

        second = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(second.status_code, 400)
        self.assertIn('hết số lần', second.data['detail'])

    def test_start_requires_assignment(self):
        other_user = User.objects.create_user(
            username='nv2', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        Employee.objects.create(tenant=self.tenant, code='NV2', name='NV2', user=other_user)
        self.client.force_authenticate(other_user)
        resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(resp.status_code, 400)
        self.assertIn('chưa được gán', resp.data['detail'])

    def test_cannot_start_unpublished_assessment(self):
        self.assessment.status = Assessment.Status.DRAFT
        self.assessment.save(update_fields=['status'])
        self.client.force_authenticate(self.learner_user)
        resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(resp.status_code, 400)


class AssessmentAssignApiTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.assessment = Assessment.objects.create(tenant=self.tenant, title='Đề demo')
        self.e1 = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', position='Phục vụ')
        self.e2 = Employee.objects.create(tenant=self.tenant, code='NV2', name='NV2', position='Thu ngân')
        self.client = APIClient()

    def test_assign_by_employee_ids(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.post(
            reverse('exam-assessment-assign', args=[self.assessment.id]),
            {'employee_ids': [self.e1.id, self.e2.id]}, format='json',
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['created'], 2)

    def test_assign_skips_already_assigned(self):
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.e1)
        self.client.force_authenticate(self.admin)
        resp = self.client.post(
            reverse('exam-assessment-assign', args=[self.assessment.id]),
            {'employee_ids': [self.e1.id, self.e2.id]}, format='json',
        )
        self.assertEqual(resp.data['created'], 1)

    def test_assign_requires_admin(self):
        trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        self.client.force_authenticate(trainer)
        resp = self.client.post(
            reverse('exam-assessment-assign', args=[self.assessment.id]),
            {'employee_ids': [self.e1.id]}, format='json',
        )
        self.assertEqual(resp.status_code, 403)


class EmployeeLearnerScopeExamsApiTests(TestCase):
    """role='employee' duoc phep goi /api/exams/ (bo sung vao ALLOWED_PREFIXES o Dot 2)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.client = APIClient()

    def test_employee_role_can_call_exams_my(self):
        self.client.force_authenticate(self.learner_user)
        resp = self.client.get(reverse('exam-my'))
        # 403 vi chua lien ket Employee, khong phai bi EmployeeLearnerScope chan truoc do
        self.assertEqual(resp.status_code, 403)
        self.assertIn('hồ sơ nhân sự', resp.data['detail'])

    def test_employee_role_blocked_from_employees_endpoint(self):
        self.client.force_authenticate(self.learner_user)
        resp = self.client.get(reverse('employee-list'))
        self.assertEqual(resp.status_code, 403)


class ExamHookTests(TestCase):
    """Dot 3 phan A/C: bat dau/nop/dat/khong dat bai thi phai sinh XapiStatement dung verb, va
    CHI dong bo ExamResult khi THAT SU dat (dung trigger cua prompt: 'graded & passed')."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.question = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Q1',
        )
        self.correct_opt = QuestionOption.objects.create(
            tenant=self.tenant, question=self.question, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(tenant=self.tenant, question=self.question, content_html='Sai')
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề 1 câu', status=Assessment.Status.PUBLISHED,
            sync_exam_type='15N', pass_mark=80,
        )
        AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=self.question)
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)
        self.client = APIClient()

    def _start(self):
        self.client.force_authenticate(self.learner_user)
        return self.client.post(reverse('exam-start', args=[self.assessment.id])).data['attempt_id']

    def test_start_logs_started_xapi(self):
        self._start()
        self.assertTrue(
            XapiStatement.objects.filter(
                employee=self.employee, verb='started', object_type='assessment', object_id=self.assessment.id,
            ).exists()
        )

    def test_pass_logs_passed_xapi_and_syncs_exam_result(self):
        attempt_id = self._start()
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': self.correct_opt.id}}, format='json',
        )
        self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))

        self.assertTrue(
            XapiStatement.objects.filter(
                employee=self.employee, verb='passed', object_type='assessment', object_id=self.assessment.id,
            ).exists()
        )
        result = ExamResult.objects.get(employee=self.employee, exam_name='15N')
        self.assertTrue(result.passed)
        self.assertEqual(result.source, 'internal_lms')

    def test_fail_logs_failed_xapi_and_does_not_sync(self):
        attempt_id = self._start()
        wrong_opt = self.question.options.exclude(id=self.correct_opt.id).first()
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': wrong_opt.id}}, format='json',
        )
        self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))

        self.assertTrue(
            XapiStatement.objects.filter(
                employee=self.employee, verb='failed', object_type='assessment', object_id=self.assessment.id,
            ).exists()
        )
        # KHONG dat -> KHONG dong bo (dung trigger 'graded & passed' cua prompt)
        self.assertFalse(ExamResult.objects.filter(employee=self.employee, exam_name='15N').exists())


class ExamEssayHookTests(TestCase):
    """Bai co essay: xAPI/dong bo CHI ban hanh khi cham tay xong (GRADED that su), khong phai
    luc nop (van con SUBMITTED cho cham)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.question = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.ESSAY, stem_html='Q1', points=10,
        )
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề tự luận', status=Assessment.Status.PUBLISHED,
            sync_exam_type='ESSAY1', pass_mark=80,
        )
        AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=self.question)
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)
        self.client = APIClient()

    def test_no_finalize_until_manual_grade_then_finalizes(self):
        self.client.force_authenticate(self.learner_user)
        attempt_id = self.client.post(reverse('exam-start', args=[self.assessment.id])).data['attempt_id']
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'text': 'Bài làm của tôi'}}, format='json',
        )
        self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))

        self.assertFalse(XapiStatement.objects.filter(employee=self.employee, verb__in=['passed', 'failed']).exists())
        self.assertFalse(ExamResult.objects.filter(employee=self.employee, exam_name='ESSAY1').exists())

        self.client.force_authenticate(self.admin)
        self.client.post(
            reverse('exam-attempt-grade', args=[attempt_id]),
            {'scores': {str(self.question.id): 9}}, format='json',
        )

        self.assertTrue(
            XapiStatement.objects.filter(
                employee=self.employee, verb='passed', object_type='assessment', object_id=self.assessment.id,
            ).exists()
        )
        self.assertTrue(ExamResult.objects.filter(employee=self.employee, exam_name='ESSAY1', passed=True).exists())


class AssessmentCustomizationApiTests(TestCase):
    """Tab 'Tuy chinh' kieu CLS (Prompt_NganHangDe_va_KyThi_kieuCLS.md muc 2): luu thoi gian lam
    bai/diem dat/so lan lam lai/che do xem lai + hien thi diem, va xac nhan review_mode/show_score
    thuc su anh huong ket qua tra ve nguoi thi (khong chi luu suong)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.question = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Q1',
        )
        self.opt_correct = QuestionOption.objects.create(
            tenant=self.tenant, question=self.question, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(tenant=self.tenant, question=self.question, content_html='Sai')
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề tùy chỉnh', status=Assessment.Status.PUBLISHED, created_by=self.admin,
        )
        AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=self.question, order=0)
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)
        self.client = APIClient()

    def test_admin_saves_customization_settings(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.patch(
            reverse('exam-assessment-detail', args=[self.assessment.id]),
            {
                'time_limit_min': 45, 'pass_mark': 70, 'max_attempts': 3,
                'questions_per_page': 5, 'show_countdown': False, 'show_score': False,
                'review_mode': 'none', 'show_grade_label': True, 'proctoring_enabled': True,
            },
            format='json',
        )
        self.assertEqual(resp.status_code, 200)
        self.assessment.refresh_from_db()
        self.assertEqual(self.assessment.time_limit_min, 45)
        self.assertEqual(self.assessment.pass_mark, 70)
        self.assertEqual(self.assessment.max_attempts, 3)
        self.assertEqual(self.assessment.questions_per_page, 5)
        self.assertFalse(self.assessment.show_countdown)
        self.assertFalse(self.assessment.show_score)
        self.assertEqual(self.assessment.review_mode, Assessment.ReviewMode.NONE)
        self.assertTrue(self.assessment.show_grade_label)
        self.assertTrue(self.assessment.proctoring_enabled)

    def _submit_and_get_result(self):
        self.client.force_authenticate(self.learner_user)
        start_resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        attempt_id = start_resp.data['attempt_id']
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': self.opt_correct.id}}, format='json',
        )
        return self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))

    def test_review_mode_none_hides_details(self):
        self.assessment.review_mode = Assessment.ReviewMode.NONE
        self.assessment.save()
        resp = self._submit_and_get_result()
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn('details', resp.data)
        self.assertIsNotNone(resp.data['score'])  # show_score mac dinh True

    def test_review_mode_full_detail_default_shows_details(self):
        resp = self._submit_and_get_result()
        self.assertIn('details', resp.data)

    def test_show_score_false_hides_score_but_keeps_passed(self):
        self.assessment.show_score = False
        self.assessment.save()
        resp = self._submit_and_get_result()
        self.assertIsNone(resp.data['score'])
        self.assertIsNone(resp.data['percent'])
        self.assertIsNotNone(resp.data['passed'])


class AssessmentManualQuestionPointsApiTests(TestCase):
    """'Tao de thu cong: chon cau hoi tu ngan hang, moi cau gan Diem' - xac nhan points_override
    dat qua API anh huong dung diem cham (khong chi luu suong)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.q1 = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Q1', points=1,
        )
        self.opt1 = QuestionOption.objects.create(
            tenant=self.tenant, question=self.q1, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(tenant=self.tenant, question=self.q1, content_html='Sai')
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề thủ công', status=Assessment.Status.PUBLISHED,
            pass_mark=50, created_by=self.admin,
        )
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)
        self.client = APIClient()

    def test_manual_pick_question_and_set_points(self):
        self.client.force_authenticate(self.admin)
        add_resp = self.client.post(
            reverse('exam-assessment-question-list'),
            {'assessment': self.assessment.id, 'question': self.q1.id, 'order': 0}, format='json',
        )
        self.assertEqual(add_resp.status_code, 201)
        aq_id = add_resp.data['id']

        patch_resp = self.client.patch(
            reverse('exam-assessment-question-detail', args=[aq_id]), {'points_override': 20}, format='json',
        )
        self.assertEqual(patch_resp.status_code, 200)
        self.assertEqual(patch_resp.data['points_override'], 20)

        self.client.force_authenticate(self.learner_user)
        start_resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        attempt_id = start_resp.data['attempt_id']
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.q1.id, 'response': {'option_id': self.opt1.id}}, format='json',
        )
        submit_resp = self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))
        # Diem cua cau (mac dinh 1) da bi ghi de thanh 20 -> dung 20 diem, khong phai 1.
        self.assertEqual(submit_resp.data['score'], Decimal('20.00'))
        self.assertEqual(submit_resp.data['max_score'], Decimal('20.00'))


class ExamSessionApiTests(TestCase):
    """Muc 3 (Ky thi): tao Ky thi giao theo vi tri -> sinh AssessmentAssignment cho nhan su khop,
    nguoi thi CHI thay bai thi trong khoang thoi gian mo, man theo doi hien dung da/chua thi."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.restaurant = Restaurant.objects.create(tenant=self.tenant, name='NH01', code='NH01')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.question = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Q1',
        )
        self.opt_correct = QuestionOption.objects.create(
            tenant=self.tenant, question=self.question, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(tenant=self.tenant, question=self.question, content_html='Sai')
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề kỳ thi', status=Assessment.Status.PUBLISHED, created_by=self.admin,
        )
        AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=self.question, order=0)

        self.matching_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.matching_employee = Employee.objects.create(
            tenant=self.tenant, code='NV1', name='Đúng vị trí', position='Phục vụ',
            restaurant=self.restaurant, user=self.matching_user,
        )
        self.other_user = User.objects.create_user(
            username='nv2', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.other_employee = Employee.objects.create(
            tenant=self.tenant, code='NV2', name='Khác vị trí', position='Thu ngân', user=self.other_user,
        )
        self.client = APIClient()

    def test_create_session_by_position_assigns_only_matching_employee(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.post(reverse('exam-session-list'), {
            'assessment': self.assessment.id, 'title': 'Kỳ thi tháng 8', 'position': 'Phục vụ',
        }, format='json')
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(resp.data['assigned_count_created'], 1)

        self.assertTrue(
            AssessmentAssignment.objects.filter(assessment=self.assessment, employee=self.matching_employee).exists()
        )
        self.assertFalse(
            AssessmentAssignment.objects.filter(assessment=self.assessment, employee=self.other_employee).exists()
        )

    def test_create_session_with_camera_supervision_enables_proctoring_and_assigns_proctors(self):
        """Nhom 3B (Prompt_Nhom3B_ThiThuViec_TuDong.md muc 3): tao Ky thi voi
        supervised_by_restaurant_camera=True phai BAT Assessment.proctoring_enabled va gan
        proctors (nguoi coi thi)."""
        trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        self.client.force_authenticate(self.admin)
        resp = self.client.post(reverse('exam-session-list'), {
            'assessment': self.assessment.id, 'title': 'Kỳ thi có giám sát', 'position': 'Phục vụ',
            'supervised_by_restaurant_camera': True, 'proctors': [trainer.id],
        }, format='json')
        self.assertEqual(resp.status_code, 201)
        self.assertTrue(resp.data['supervised_by_restaurant_camera'])
        self.assertEqual(resp.data['proctors_detail'][0]['id'], trainer.id)

        self.assessment.refresh_from_db()
        self.assertTrue(self.assessment.proctoring_enabled)
        session = ExamSession.objects.get(pk=resp.data['id'])
        self.assertIn(trainer, session.proctors.all())

    def test_create_session_requires_admin(self):
        self.client.force_authenticate(self.matching_user)
        resp = self.client.post(reverse('exam-session-list'), {
            'assessment': self.assessment.id, 'position': 'Phục vụ',
        }, format='json')
        self.assertEqual(resp.status_code, 403)

    def test_create_session_requires_target(self):
        self.client.force_authenticate(self.admin)
        resp = self.client.post(reverse('exam-session-list'), {
            'assessment': self.assessment.id, 'title': 'Không có mục tiêu',
        }, format='json')
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(ExamSession.objects.exists())  # rollback dung, khong bo lai session mo côi

    def test_employee_only_sees_exam_within_open_window(self):
        now = timezone.now()
        session = ExamSession.objects.create(
            tenant=self.tenant, title='Kỳ thi tương lai', assessment=self.assessment,
            start_at=now + timezone.timedelta(days=1), end_at=now + timezone.timedelta(days=2),
            created_by=self.admin,
        )
        AssessmentAssignment.objects.create(
            tenant=self.tenant, assessment=self.assessment, employee=self.matching_employee, exam_session=session,
        )
        self.client.force_authenticate(self.matching_user)
        resp = self.client.get(reverse('exam-my'))
        self.assertEqual(resp.data, [])  # chua mo -> khong thay

        session.start_at = now - timezone.timedelta(hours=1)
        session.end_at = now + timezone.timedelta(hours=1)
        session.save()
        resp = self.client.get(reverse('exam-my'))
        self.assertEqual(len(resp.data), 1)  # dang mo -> thay

        session.start_at = now - timezone.timedelta(days=2)
        session.end_at = now - timezone.timedelta(days=1)
        session.save()
        resp = self.client.get(reverse('exam-my'))
        self.assertEqual(resp.data, [])  # da dong -> khong con thay

    def test_start_attempt_blocked_outside_session_window(self):
        now = timezone.now()
        session = ExamSession.objects.create(
            tenant=self.tenant, title='Kỳ thi đã đóng', assessment=self.assessment,
            start_at=now - timezone.timedelta(days=2), end_at=now - timezone.timedelta(days=1),
            created_by=self.admin,
        )
        AssessmentAssignment.objects.create(
            tenant=self.tenant, assessment=self.assessment, employee=self.matching_employee, exam_session=session,
        )
        self.client.force_authenticate(self.matching_user)
        resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(resp.status_code, 400)
        self.assertIn('đóng', resp.data['detail'])

    def test_manual_assignment_without_session_is_unaffected(self):
        """Gan tay truc tiep (Dot 2, khong qua Ky thi) van hien/lam bai duoc binh thuong, khong
        bi anh huong boi tinh nang gioi han thoi gian moi."""
        AssessmentAssignment.objects.create(
            tenant=self.tenant, assessment=self.assessment, employee=self.matching_employee,
        )
        self.client.force_authenticate(self.matching_user)
        resp = self.client.get(reverse('exam-my'))
        self.assertEqual(len(resp.data), 1)
        start_resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(start_resp.status_code, 200)

    def test_tracking_shows_done_and_not_done(self):
        session = ExamSession.objects.create(
            tenant=self.tenant, title='Kỳ thi theo dõi', assessment=self.assessment, created_by=self.admin,
        )
        AssessmentAssignment.objects.create(
            tenant=self.tenant, assessment=self.assessment, employee=self.matching_employee, exam_session=session,
        )
        AssessmentAssignment.objects.create(
            tenant=self.tenant, assessment=self.assessment, employee=self.other_employee, exam_session=session,
        )

        self.client.force_authenticate(self.matching_user)
        start_resp = self.client.post(reverse('exam-start', args=[self.assessment.id]))
        attempt_id = start_resp.data['attempt_id']
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': self.opt_correct.id}}, format='json',
        )
        self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))

        self.client.force_authenticate(self.admin)
        resp = self.client.get(reverse('exam-session-tracking', args=[session.id]))
        self.assertEqual(resp.status_code, 200)
        by_code = {r['employee_code']: r for r in resp.data}
        self.assertTrue(by_code['NV1']['done'])
        self.assertTrue(by_code['NV1']['passed'])
        self.assertFalse(by_code['NV2']['done'])
        self.assertIsNone(by_code['NV2']['passed'])

    def test_tracking_requires_admin(self):
        session = ExamSession.objects.create(
            tenant=self.tenant, title='X', assessment=self.assessment, created_by=self.admin,
        )
        self.client.force_authenticate(self.matching_user)
        resp = self.client.get(reverse('exam-session-tracking', args=[session.id]))
        self.assertEqual(resp.status_code, 403)


# ==================================================================== import_cls_questions

CLS_HEADER = [
    '      STT (*)      ', '      STT Câu Hỏi Cha      ', '      Chủ Đề (*)      ',
    '      Kiểu Câu Hỏi (*)      ', '      Nội Dung (*)      ', '      Mã tác giả      ',
    '      Tác giả      ', '      Đường Dẫn Tệp Tin      ', '      Đáp Án Đúng      ',
    '      Giải Thích Kết Quả      ', '      Xáo Trộn      ', '      Câu trả lời không xáo trộn      ',
    '      Cấp Độ (*)      ', '      Câu Trả Lời 1      ', '      Câu Trả Lời 2      ',
    '      Câu Trả Lời 3      ', '      Câu Trả Lời 4      ', '      Câu Trả Lời 5      ',
    '      Câu Trả Lời 6      ', '      Câu Trả Lời 7      ', '      Câu Trả Lời 8      ',
]


def _cls_row(
    stt, topic, qtype, content, correct, level, answers,
    explanation='', media='', author_code='', author='',
):
    answers = (answers + [''] * 8)[:8]
    return [
        stt, '', topic, qtype, content, author_code, author, media, correct, explanation,
        'Không', '1, 2, 3, 4', level, *answers,
    ]


def _build_cls_workbook(rows, sheet_name='Câu Hỏi', header=None):
    """Dung dinh dang GIONG HET file that (dong 1 = tieu de gop o, dong 2 = header co khoang
    trang thua can .strip(), dong 3+ = du lieu) - xem exams/cls_import.py."""
    import io as _io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['MẪU NGÂN HÀNG CÂU HỎI'])
    ws.append(header if header is not None else CLS_HEADER)
    for row in rows:
        ws.append(row)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ClsImportParseTests(TestCase):
    """parse_workbook() thuan tuy (khong dong DB) - dung workbook dung dinh dang that."""

    def test_parses_single_choice_correct_option(self):
        from exams.cls_import import parse_workbook

        wb = _build_cls_workbook([
            _cls_row('1', 'Chủ đề A', 'Trắc nghiệm một lựa chọn', 'Câu hỏi 1?', '2', 'Nhận biết',
                     ['Sai 1', 'Đúng', 'Sai 2', 'Sai 3']),
        ])
        result = parse_workbook(wb)
        self.assertEqual(len(result['parsed']), 1)
        self.assertEqual(len(result['skipped']), 0)
        row = result['parsed'][0]
        self.assertEqual(row['type'], Question.Type.SINGLE)
        self.assertEqual(row['bank_name'], 'Chủ đề A')
        self.assertEqual(row['difficulty'], Question.Difficulty.EASY)
        self.assertEqual([o['text'] for o in row['options']], ['Sai 1', 'Đúng', 'Sai 2', 'Sai 3'])
        self.assertEqual([o['is_correct'] for o in row['options']], [False, True, False, False])

    def test_parses_multiple_choice_correct_options(self):
        from exams.cls_import import parse_workbook

        wb = _build_cls_workbook([
            _cls_row('2', 'Chủ đề A', 'Trắc nghiệm nhiều lựa chọn', 'Câu hỏi nhiều đáp án?', '1, 3', 'Vận Dụng',
                     ['Đúng 1', 'Sai 1', 'Đúng 2', 'Sai 2']),
        ])
        result = parse_workbook(wb)
        row = result['parsed'][0]
        self.assertEqual(row['type'], Question.Type.MULTIPLE)
        self.assertEqual(row['difficulty'], Question.Difficulty.MEDIUM)
        self.assertEqual([o['is_correct'] for o in row['options']], [True, False, True, False])

    def test_skips_row_missing_content(self):
        from exams.cls_import import parse_workbook

        wb = _build_cls_workbook([
            _cls_row('3', 'Chủ đề A', 'Trắc nghiệm một lựa chọn', '', '1', 'Nhận biết', ['A', 'B']),
        ])
        result = parse_workbook(wb)
        self.assertEqual(result['parsed'], [])
        self.assertEqual(len(result['skipped']), 1)
        self.assertIn('Nội Dung', result['skipped'][0]['reason'])

    def test_skips_unsupported_question_type(self):
        from exams.cls_import import parse_workbook

        wb = _build_cls_workbook([
            _cls_row('4', 'Chủ đề A', 'Gạch chân', 'Câu hỏi dạng lạ?', '1', 'Nhận biết', ['A', 'B']),
        ])
        result = parse_workbook(wb)
        self.assertEqual(result['parsed'], [])
        self.assertEqual(len(result['skipped']), 1)
        self.assertIn('Kiểu câu hỏi không hỗ trợ', result['skipped'][0]['reason'])

    def test_skips_row_without_valid_correct_answer(self):
        from exams.cls_import import parse_workbook

        wb = _build_cls_workbook([
            _cls_row('5', 'Chủ đề A', 'Trắc nghiệm một lựa chọn', 'Câu hỏi thiếu đáp án đúng?', '',
                     'Nhận biết', ['A', 'B']),
        ])
        result = parse_workbook(wb)
        self.assertEqual(result['parsed'], [])
        self.assertEqual(len(result['skipped']), 1)
        self.assertIn('Đáp Án Đúng', result['skipped'][0]['reason'])

    def test_finds_header_row_even_without_leading_title_row(self):
        """Phong ho file chu de khac co the KHONG co dong tieu de gop o rieng (header ngay dong
        1) - _find_header_row phai van nhan dung, khong gia dinh cung dong 2."""
        from exams.cls_import import parse_workbook
        import io as _io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Câu Hỏi'
        ws.append(CLS_HEADER)
        ws.append(_cls_row('1', 'Chủ đề B', 'Trắc nghiệm một lựa chọn', 'Câu hỏi?', '1', 'Nhận biết', ['A', 'B']))
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        result = parse_workbook(buf)
        self.assertEqual(len(result['parsed']), 1)


class ClsImportWriteTests(TestCase):
    """import_rows() - ghi DB (hoac chi dem neu dry_run)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')

    def _sample_rows(self):
        from exams.cls_import import parse_workbook

        wb = _build_cls_workbook([
            _cls_row('1', 'Câu hỏi thi Phục vụ', 'Trắc nghiệm một lựa chọn', 'Câu 1?', '1', 'Nhận biết',
                     ['Đúng', 'Sai']),
            _cls_row('2', 'Câu hỏi thi Phục vụ', 'Trắc nghiệm nhiều lựa chọn', 'Câu 2?', '1, 2', 'Vận Dụng',
                     ['Đúng 1', 'Đúng 2', 'Sai']),
        ])
        return parse_workbook(wb)['parsed']

    def test_dry_run_reports_stats_without_writing_anything(self):
        from exams.cls_import import import_rows

        stats = import_rows(self.tenant, self._sample_rows(), dry_run=True)
        self.assertEqual(stats['banks_created'], 1)
        self.assertEqual(stats['questions_created'], 2)
        self.assertEqual(stats['single_created'], 1)
        self.assertEqual(stats['multiple_created'], 1)
        self.assertEqual(stats['options_created'], 5)
        self.assertEqual(QuestionBank.objects.count(), 0)
        self.assertEqual(Question.objects.count(), 0)

    def test_real_run_creates_bank_question_options(self):
        from exams.cls_import import import_rows

        stats = import_rows(self.tenant, self._sample_rows(), dry_run=False)
        self.assertEqual(stats['questions_created'], 2)
        bank = QuestionBank.objects.get(tenant=self.tenant, name='Câu hỏi thi Phục vụ')
        self.assertEqual(Question.objects.filter(bank=bank).count(), 2)
        q1 = Question.objects.get(bank=bank, stem_html='Câu 1?')
        self.assertEqual(q1.type, Question.Type.SINGLE)
        self.assertTrue(q1.options.get(content_html='Đúng').is_correct)
        self.assertFalse(q1.options.get(content_html='Sai').is_correct)

    def test_running_twice_is_idempotent(self):
        from exams.cls_import import import_rows

        rows = self._sample_rows()
        import_rows(self.tenant, rows, dry_run=False)
        self.assertEqual(QuestionBank.objects.count(), 1)
        self.assertEqual(Question.objects.count(), 2)
        self.assertEqual(QuestionOption.objects.count(), 5)

        stats2 = import_rows(self.tenant, rows, dry_run=False)
        self.assertEqual(stats2['questions_created'], 0)
        self.assertEqual(stats2['questions_skipped_duplicate'], 2)
        self.assertEqual(QuestionBank.objects.count(), 1)
        self.assertEqual(Question.objects.count(), 2)
        self.assertEqual(QuestionOption.objects.count(), 5)


class ImportClsQuestionsCommandTests(TestCase):
    """Test o muc lenh CLI (call_command) - tenant khong ton tai, file khong ton tai, dry-run
    that qua tham so --file/--dry-run."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')

    def _write_temp_xlsx(self):
        import tempfile

        wb = _build_cls_workbook([
            _cls_row('1', 'Chủ đề CLI', 'Trắc nghiệm một lựa chọn', 'Câu CLI?', '1', 'Nhận biết', ['Đúng', 'Sai']),
        ])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(wb.read())
        tmp.close()
        return tmp.name

    def test_dry_run_via_command_does_not_write(self):
        from django.core.management import call_command

        path = self._write_temp_xlsx()
        call_command('import_cls_questions', file=path, tenant='Demo Tenant', dry_run=True)
        self.assertEqual(QuestionBank.objects.count(), 0)

    def test_real_run_via_command_writes(self):
        from django.core.management import call_command

        path = self._write_temp_xlsx()
        call_command('import_cls_questions', file=path, tenant='Demo Tenant')
        self.assertEqual(QuestionBank.objects.filter(name='Chủ đề CLI').count(), 1)
        self.assertEqual(Question.objects.filter(stem_html='Câu CLI?').count(), 1)

    def test_missing_tenant_raises(self):
        from django.core.management import CommandError, call_command

        path = self._write_temp_xlsx()
        with self.assertRaises(CommandError):
            call_command('import_cls_questions', file=path, tenant='Tenant không tồn tại')

    def test_missing_file_raises(self):
        from django.core.management import CommandError, call_command

        with self.assertRaises(CommandError):
            call_command('import_cls_questions', file='/khong/ton/tai.xlsx', tenant='Demo Tenant')

    def test_cls_file_with_optional_nang_luc_column_sets_competency(self):
        """Muc 3 cua Prompt_GanNangLuc_CauHoi_Excel.md: cot 'Nang luc' TUY CHON trong file CLS -
        co thi gan luon, khop ten khong khop thi bo trong (khong chan tao cau hoi)."""
        from exams.cls_import import import_rows, parse_workbook

        group = CompetencyGroup.objects.create(tenant=self.tenant, code='A1', name='Nhóm A1')
        comp = Competency.objects.create(tenant=self.tenant, group=group, name='Xử lý order & POS')

        header_with_nl = CLS_HEADER + ['      Năng lực      ']
        wb = _build_cls_workbook([
            [*_cls_row('1', 'Chủ đề C', 'Trắc nghiệm một lựa chọn', 'Câu 1?', '1', 'Nhận biết', ['A', 'B']),
             'Xử lý order & POS'],
            [*_cls_row('2', 'Chủ đề C', 'Trắc nghiệm một lựa chọn', 'Câu 2?', '1', 'Nhận biết', ['A', 'B']),
             'Năng lực không tồn tại'],
        ], header=header_with_nl)

        parsed = parse_workbook(wb)['parsed']
        stats = import_rows(self.tenant, parsed, dry_run=False)
        self.assertEqual(stats['competency_matched'], 1)
        self.assertEqual(stats['competency_unmatched'], 1)

        q1 = Question.objects.get(tenant=self.tenant, stem_html='Câu 1?')
        q2 = Question.objects.get(tenant=self.tenant, stem_html='Câu 2?')
        self.assertEqual(q1.competency_id, comp.id)
        self.assertIsNone(q2.competency_id)


def _build_competency_import_workbook(rows):
    """rows: [(question_id, competency_text), ...] - dinh dang GIONG HET file xuat boi
    export_workbook (xem exams/competency_assign.py)."""
    import io as _io

    import openpyxl

    from exams.competency_assign import HEADER

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gan nang luc'
    ws.append(HEADER)
    for question_id, competency_text in rows:
        ws.append([question_id, f'Nội dung câu {question_id}', 'Chủ đề X', 'Trắc nghiệm một lựa chọn', competency_text])
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class CompetencyExportImportTests(TestCase):
    """Prompt_GanNangLuc_CauHoi_Excel.md: xuat/nhap Excel gan nang luc cho cau hoi."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.group = CompetencyGroup.objects.create(tenant=self.tenant, code='A1', name='Nhóm A1')
        self.comp_a = Competency.objects.create(tenant=self.tenant, group=self.group, name='Xử lý order & POS', order=1)
        self.comp_b = Competency.objects.create(
            tenant=self.tenant, group=self.group, name='Quy trình phục vụ chuẩn', order=2,
        )
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Câu hỏi thi Phục vụ Kampong')
        self.q1 = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Câu hỏi 1?',
        )
        self.q2 = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Câu hỏi 2?',
            competency=self.comp_a,
        )
        self.client = APIClient()

    # ---- export_workbook ----

    def test_export_workbook_has_data_and_catalog_sheets_with_dropdown(self):
        from exams.competency_assign import SHEET_CATALOG, SHEET_DATA, export_workbook

        wb = export_workbook(self.tenant, Question.objects.filter(tenant=self.tenant, bank=self.bank))
        self.assertIn(SHEET_DATA, wb.sheetnames)
        self.assertIn(SHEET_CATALOG, wb.sheetnames)

        ws = wb[SHEET_DATA]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('Mã câu hỏi', 'Nội dung câu hỏi', 'Chủ đề', 'Dạng', 'NĂNG LỰC (chọn)'))
        by_id = {r[0]: r for r in rows[1:]}
        self.assertEqual(by_id[self.q1.id][4], '')  # chua gan -> de trong
        self.assertEqual(by_id[self.q2.id][4], 'Xử lý order & POS')  # da gan -> dien san

        catalog_names = [r[0] for r in ws.parent[SHEET_CATALOG].iter_rows(values_only=True)][1:]
        self.assertEqual(set(catalog_names), {'Xử lý order & POS', 'Quy trình phục vụ chuẩn'})

        dv = ws.data_validations.dataValidation[0]
        self.assertEqual(dv.type, 'list')
        self.assertIn(SHEET_CATALOG, dv.formula1)

    def test_export_endpoint_requires_admin_and_respects_filter(self):
        trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        self.client.force_authenticate(trainer)
        resp = self.client.get(reverse('exam-question-export-competency'))
        self.assertEqual(resp.status_code, 403)

        self.client.force_authenticate(self.admin)
        resp = self.client.get(reverse('exam-question-export-competency'), {'bank': self.bank.id})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ---- import: apply_competency_assignments ----

    def test_matches_by_id_and_sets_competency(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(self.q1.id, 'Xử lý order & POS')])
        raw_rows = parse_import_workbook(wb)
        result = apply_competency_assignments(self.tenant, raw_rows, dry_run=False)
        self.assertEqual(result['stats']['will_assign'], 1)
        self.assertEqual(result['errors'], [])
        self.q1.refresh_from_db()
        self.assertEqual(self.q1.competency_id, self.comp_a.id)

    def test_name_matching_is_diacritics_and_whitespace_insensitive(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(self.q1.id, '  xu ly order & pos  ')])
        raw_rows = parse_import_workbook(wb)
        result = apply_competency_assignments(self.tenant, raw_rows, dry_run=False)
        self.assertEqual(result['stats']['will_assign'], 1)
        self.q1.refresh_from_db()
        self.assertEqual(self.q1.competency_id, self.comp_a.id)

    def test_unknown_competency_name_reports_error_without_creating(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(self.q1.id, 'Năng lực không có thật')])
        raw_rows = parse_import_workbook(wb)
        result = apply_competency_assignments(self.tenant, raw_rows, dry_run=False)
        self.assertEqual(result['stats']['errors'], 1)
        self.assertIn('Không khớp', result['errors'][0]['reason'])
        self.q1.refresh_from_db()
        self.assertIsNone(self.q1.competency_id)
        self.assertEqual(Competency.objects.count(), 2)  # khong tao them nang luc moi

    def test_blank_competency_cell_keeps_existing_value(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(self.q2.id, '')])
        raw_rows = parse_import_workbook(wb)
        result = apply_competency_assignments(self.tenant, raw_rows, dry_run=False)
        self.assertEqual(result['stats']['unchanged_blank'], 1)
        self.q2.refresh_from_db()
        self.assertEqual(self.q2.competency_id, self.comp_a.id)  # giu nguyen, khong bi xoa

    def test_unknown_question_id_reports_error(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(999999, 'Xử lý order & POS')])
        raw_rows = parse_import_workbook(wb)
        result = apply_competency_assignments(self.tenant, raw_rows, dry_run=False)
        self.assertEqual(result['stats']['errors'], 1)
        self.assertIn('Không tìm thấy câu hỏi', result['errors'][0]['reason'])

    def test_dry_run_does_not_write(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(self.q1.id, 'Xử lý order & POS')])
        raw_rows = parse_import_workbook(wb)
        result = apply_competency_assignments(self.tenant, raw_rows, dry_run=True)
        self.assertEqual(result['stats']['will_assign'], 1)
        self.q1.refresh_from_db()
        self.assertIsNone(self.q1.competency_id)

    def test_running_twice_is_idempotent(self):
        from exams.competency_assign import apply_competency_assignments, parse_import_workbook

        wb = _build_competency_import_workbook([(self.q1.id, 'Xử lý order & POS')])
        apply_competency_assignments(self.tenant, parse_import_workbook(wb), dry_run=False)
        self.q1.refresh_from_db()
        self.assertEqual(self.q1.competency_id, self.comp_a.id)

        wb2 = _build_competency_import_workbook([(self.q1.id, 'Xử lý order & POS')])
        result2 = apply_competency_assignments(self.tenant, parse_import_workbook(wb2), dry_run=False)
        self.assertEqual(result2['stats']['will_assign'], 1)
        self.assertEqual(result2['errors'], [])
        self.q1.refresh_from_db()
        self.assertEqual(self.q1.competency_id, self.comp_a.id)  # van dung, khong doi/loi gi them

    # ---- import endpoint (preview -> confirm) ----

    def test_import_endpoint_preview_then_confirm(self):
        wb = _build_competency_import_workbook([(self.q1.id, 'Xử lý order & POS')])
        self.client.force_authenticate(self.admin)

        from django.core.files.uploadedfile import SimpleUploadedFile

        upload = SimpleUploadedFile(
            'gan_nang_luc.xlsx', wb.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp = self.client.post(
            reverse('exam-question-import-competency'), {'file': upload}, format='multipart',
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.data['dry_run'])
        self.assertEqual(resp.data['stats']['will_assign'], 1)
        self.q1.refresh_from_db()
        self.assertIsNone(self.q1.competency_id)  # preview - chua ghi

        wb2 = _build_competency_import_workbook([(self.q1.id, 'Xử lý order & POS')])
        upload2 = SimpleUploadedFile(
            'gan_nang_luc.xlsx', wb2.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp2 = self.client.post(
            reverse('exam-question-import-competency'), {'file': upload2, 'dry_run': 'false'}, format='multipart',
        )
        self.assertEqual(resp2.status_code, 200)
        self.assertFalse(resp2.data['dry_run'])
        self.q1.refresh_from_db()
        self.assertEqual(self.q1.competency_id, self.comp_a.id)

    def test_import_endpoint_requires_admin(self):
        trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        self.client.force_authenticate(trainer)
        resp = self.client.post(reverse('exam-question-import-competency'), {}, format='multipart')
        self.assertEqual(resp.status_code, 403)


# ==================================================================== Giai doan A: Proctoring


class ProctoringBaseTestCase(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.admin = User.objects.create_user(username='admin1', password='x', tenant=self.tenant, role='admin')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.question = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Q1',
        )
        self.opt_correct = QuestionOption.objects.create(
            tenant=self.tenant, question=self.question, content_html='Đúng', is_correct=True,
        )
        QuestionOption.objects.create(tenant=self.tenant, question=self.question, content_html='Sai')
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề có giám sát', status=Assessment.Status.PUBLISHED, created_by=self.admin,
        )
        AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=self.question, order=0)
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)
        self.client = APIClient()

    def _start(self, password=None):
        self.client.force_authenticate(self.learner_user)
        payload = {'password': password} if password is not None else {}
        return self.client.post(reverse('exam-start', args=[self.assessment.id]), payload, format='json')


class A1AccessPasswordTests(ProctoringBaseTestCase):
    """A1: mat khau vao de (tuy chon)."""

    def test_no_password_set_starts_normally(self):
        resp = self._start()
        self.assertEqual(resp.status_code, 200)

    def test_wrong_password_rejected(self):
        self.assessment.access_password = 'bimat123'
        self.assessment.save()
        resp = self._start(password='sai')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('mật khẩu', resp.data['detail'])
        self.assertFalse(Attempt.objects.exists())

    def test_missing_password_rejected_when_required(self):
        self.assessment.access_password = 'bimat123'
        self.assessment.save()
        resp = self._start()
        self.assertEqual(resp.status_code, 400)

    def test_correct_password_starts(self):
        self.assessment.access_password = 'bimat123'
        self.assessment.save()
        resp = self._start(password='bimat123')
        self.assertEqual(resp.status_code, 200)

    def test_access_password_never_leaks_in_api_response_but_has_password_does(self):
        self.client.force_authenticate(self.admin)
        patch_resp = self.client.patch(
            reverse('exam-assessment-detail', args=[self.assessment.id]), {'access_password': 'bimat123'}, format='json',
        )
        self.assertEqual(patch_resp.status_code, 200)
        self.assertNotIn('access_password', patch_resp.data)
        self.assertTrue(patch_resp.data['has_password'])

        get_resp = self.client.get(reverse('exam-assessment-detail', args=[self.assessment.id]))
        self.assertNotIn('access_password', get_resp.data)
        self.assertTrue(get_resp.data['has_password'])
        self.assessment.refresh_from_db()
        self.assertEqual(self.assessment.access_password, 'bimat123')


class A1TimeLimitEnforcementTests(ProctoringBaseTestCase):
    """A1: 'gioi han thoi gian' - bo sung enforcement server-side (truoc day chi la dem nguoc FE)."""

    def test_no_time_limit_unaffected(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        save_resp = self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': self.opt_correct.id}}, format='json',
        )
        self.assertEqual(save_resp.status_code, 200)

    def test_answers_rejected_and_auto_submitted_after_time_limit(self):
        self.assessment.time_limit_min = 10
        self.assessment.save()
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        attempt = Attempt.objects.get(pk=attempt_id)
        attempt.started_at = timezone.now() - timedelta(minutes=11)
        attempt.save(update_fields=['started_at'])

        save_resp = self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': self.opt_correct.id}}, format='json',
        )
        self.assertEqual(save_resp.status_code, 400)
        self.assertIn('hết thời gian', save_resp.data['detail'])

        attempt.refresh_from_db()
        self.assertEqual(attempt.status, Attempt.Status.GRADED)
        self.assertIsNotNone(attempt.submitted_at)

    def test_answers_ok_before_time_limit(self):
        self.assessment.time_limit_min = 30
        self.assessment.save()
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        save_resp = self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': self.opt_correct.id}}, format='json',
        )
        self.assertEqual(save_resp.status_code, 200)


class A2A3ProctoringEventApiTests(ProctoringBaseTestCase):
    """A2 (snapshot) + A3 (roi tab / tu nop sau N lan)."""

    def setUp(self):
        super().setUp()
        self.assessment.proctoring_enabled = True
        self.assessment.save()

    def test_attempt_detail_includes_proctoring_config_only_when_enabled(self):
        self.assessment.max_attempts = 2
        self.assessment.save()

        resp = self._start()
        self.assertIsNotNone(resp.data['proctoring'])
        self.assertTrue(resp.data['proctoring']['enabled'])
        self.assertEqual(resp.data['proctoring']['snapshot_interval_sec'], 45)
        self.client.post(reverse('exam-attempt-submit', args=[resp.data['attempt_id']]))

        self.assessment.proctoring_enabled = False
        self.assessment.save()
        resp2 = self._start()
        self.assertIsNone(resp2.data['proctoring'])

    def test_log_tab_leave_event(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        ev_resp = self.client.post(
            reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'tab_leave'}, format='json',
        )
        self.assertEqual(ev_resp.status_code, 200)
        self.assertFalse(ev_resp.data['auto_submitted'])
        self.assertEqual(ProctoringEvent.objects.filter(attempt_id=attempt_id, type='tab_leave').count(), 1)

    def test_invalid_event_type_rejected(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        ev_resp = self.client.post(
            reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'not_a_type'}, format='json',
        )
        self.assertEqual(ev_resp.status_code, 400)

    def test_cannot_log_event_for_other_employees_attempt(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        other_user = User.objects.create_user(
            username='nv2', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        Employee.objects.create(tenant=self.tenant, code='NV2', name='NV2', user=other_user)
        self.client.force_authenticate(other_user)
        ev_resp = self.client.post(
            reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'tab_leave'}, format='json',
        )
        self.assertEqual(ev_resp.status_code, 404)

    def test_auto_submit_after_tab_leave_limit_reached(self):
        self.assessment.tab_leave_auto_submit_limit = 3
        self.assessment.save()
        resp = self._start()
        attempt_id = resp.data['attempt_id']

        for _ in range(2):
            ev_resp = self.client.post(
                reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'tab_leave'}, format='json',
            )
            self.assertFalse(ev_resp.data['auto_submitted'])

        final_resp = self.client.post(
            reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'tab_leave'}, format='json',
        )
        self.assertTrue(final_resp.data['auto_submitted'])
        self.assertIn('result', final_resp.data)

        attempt = Attempt.objects.get(pk=attempt_id)
        self.assertEqual(attempt.status, Attempt.Status.GRADED)

        # Da nop roi - roi tab them nua khong con tu nop lai (khong loi, chi ghi log binh thuong).
        after_resp = self.client.post(
            reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'tab_leave'}, format='json',
        )
        self.assertEqual(after_resp.status_code, 200)
        self.assertFalse(after_resp.data['auto_submitted'])

    def test_no_auto_submit_when_limit_not_configured(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        for _ in range(10):
            ev_resp = self.client.post(
                reverse('exam-attempt-proctoring-event', args=[attempt_id]), {'type': 'tab_leave'}, format='json',
            )
            self.assertFalse(ev_resp.data['auto_submitted'])
        attempt = Attempt.objects.get(pk=attempt_id)
        self.assertEqual(attempt.status, Attempt.Status.IN_PROGRESS)

    @patch('checklist.storage.upload_data_url', return_value='https://pub-x.r2.dev/proctoring/snap.jpg')
    def test_snapshot_event_uploads_image(self, mock_upload):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        ev_resp = self.client.post(
            reverse('exam-attempt-proctoring-event', args=[attempt_id]),
            {'type': 'snapshot', 'image': 'data:image/jpeg;base64,AAAA'}, format='json',
        )
        self.assertEqual(ev_resp.status_code, 200)
        mock_upload.assert_called_once()
        event = ProctoringEvent.objects.get(pk=ev_resp.data['event_id'])
        self.assertEqual(event.image_url, 'https://pub-x.r2.dev/proctoring/snap.jpg')

    def test_snapshot_upload_failure_returns_400(self):
        from checklist.storage import StorageError

        with patch('checklist.storage.upload_data_url', side_effect=StorageError('lỗi upload')):
            resp = self._start()
            attempt_id = resp.data['attempt_id']
            ev_resp = self.client.post(
                reverse('exam-attempt-proctoring-event', args=[attempt_id]),
                {'type': 'snapshot', 'image': 'data:image/jpeg;base64,AAAA'}, format='json',
            )
            self.assertEqual(ev_resp.status_code, 400)


class A4EvidenceViewingTests(ProctoringBaseTestCase):
    """A4: man giam khao xem bang chung + dat co nghi van."""

    def setUp(self):
        super().setUp()
        self.assessment.proctoring_enabled = True
        self.assessment.save()

    def test_evaluator_sees_timeline_and_suspicion_score(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        attempt = Attempt.objects.get(pk=attempt_id)
        ProctoringEvent.objects.create(tenant=self.tenant, attempt=attempt, type='tab_leave')
        ProctoringEvent.objects.create(tenant=self.tenant, attempt=attempt, type='no_face')
        ProctoringEvent.objects.create(
            tenant=self.tenant, attempt=attempt, type='snapshot', image_url='https://pub-x.r2.dev/a.jpg',
        )

        self.client.force_authenticate(self.admin)
        timeline_resp = self.client.get(reverse('exam-attempt-proctoring', args=[attempt_id]))
        self.assertEqual(timeline_resp.status_code, 200)
        self.assertEqual(len(timeline_resp.data['events']), 3)
        self.assertEqual(timeline_resp.data['counts']['tab_leave'], 1)
        self.assertEqual(timeline_resp.data['counts']['no_face'], 1)
        self.assertEqual(timeline_resp.data['suspicion_score'], 2)  # tab_leave + no_face, khong tinh snapshot
        snapshot_events = [e for e in timeline_resp.data['events'] if e['type'] == 'snapshot']
        self.assertEqual(snapshot_events[0]['image_url'], 'https://pub-x.r2.dev/a.jpg')

    def test_timeline_exposes_proctors_and_camera_supervision_flag(self):
        """Nhom 3B muc 3: doi chieu bang chung - man Cham bai can biet ai duoc phan cong coi
        thi + co dang giam sat camera nha hang hay khong, doc tu ExamSession cua assignment."""
        trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        session = ExamSession.objects.create(
            tenant=self.tenant, title='Thi thử việc', assessment=self.assessment,
            supervised_by_restaurant_camera=True,
        )
        session.proctors.add(trainer)
        assignment = AssessmentAssignment.objects.get(assessment=self.assessment, employee=self.employee)
        assignment.exam_session = session
        assignment.save(update_fields=['exam_session'])

        resp = self._start()
        attempt_id = resp.data['attempt_id']
        self.client.force_authenticate(self.admin)
        timeline_resp = self.client.get(reverse('exam-attempt-proctoring', args=[attempt_id]))
        self.assertEqual(timeline_resp.status_code, 200)
        self.assertTrue(timeline_resp.data['supervised_by_restaurant_camera'])
        self.assertEqual(timeline_resp.data['proctors'][0]['id'], trainer.id)

    def test_timeline_no_session_means_no_camera_supervision(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        self.client.force_authenticate(self.admin)
        timeline_resp = self.client.get(reverse('exam-attempt-proctoring', args=[attempt_id]))
        self.assertFalse(timeline_resp.data['supervised_by_restaurant_camera'])
        self.assertEqual(timeline_resp.data['proctors'], [])

    def test_non_evaluator_cannot_view_timeline(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        trainer = User.objects.create_user(username='trainer1', password='x', tenant=self.tenant, role='trainer')
        self.client.force_authenticate(trainer)
        timeline_resp = self.client.get(reverse('exam-attempt-proctoring', args=[attempt_id]))
        self.assertEqual(timeline_resp.status_code, 403)

    def test_evaluator_can_flag_and_unflag_attempt(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        self.client.force_authenticate(self.admin)

        flag_resp = self.client.post(reverse('exam-attempt-flag', args=[attempt_id]), {'flagged': True}, format='json')
        self.assertEqual(flag_resp.status_code, 200)
        self.assertTrue(flag_resp.data['flagged_suspicious'])
        self.assertTrue(Attempt.objects.get(pk=attempt_id).flagged_suspicious)

        unflag_resp = self.client.post(reverse('exam-attempt-flag', args=[attempt_id]), {'flagged': False}, format='json')
        self.assertFalse(unflag_resp.data['flagged_suspicious'])

    def test_results_list_includes_flag_and_event_count(self):
        resp = self._start()
        attempt_id = resp.data['attempt_id']
        attempt = Attempt.objects.get(pk=attempt_id)
        ProctoringEvent.objects.create(tenant=self.tenant, attempt=attempt, type='tab_leave')

        self.client.force_authenticate(self.admin)
        results_resp = self.client.get(reverse('exam-assessment-results', args=[self.assessment.id]))
        self.assertEqual(results_resp.status_code, 200)
        row = next(r for r in results_resp.data if r['id'] == attempt_id)
        self.assertEqual(row['proctoring_event_count'], 1)
        self.assertFalse(row['flagged_suspicious'])


class A1BasicControlsRegressionTests(ProctoringBaseTestCase):
    """A1: xac nhan tron cau/tron dap an/gioi han so lan DA hoat dong tu truoc (khong can sua) -
    test nay chi la XAC NHAN (nghiem thu 'kiem tra dang hoat dong'), khong test logic moi."""

    def test_max_attempts_still_enforced(self):
        self.assessment.max_attempts = 1
        self.assessment.save()
        self._start()
        second_resp = self._start()
        self.assertEqual(second_resp.status_code, 400)
        self.assertIn('hết số lần', second_resp.data['detail'])


class EmployeeAttemptsReviewTests(TestCase):
    """Nhom 1 muc B.3 (Prompt_Nhom1_NhanSu_NguoiDung.md): exams.services.employee_attempts_review
    - lich su lam bai KEM chi tiet dung/sai + dap an da chon tung cau, dung cho man Chi tiet
    nhan su (KHONG bi gioi han boi show_result_mode/review_mode cua de, khac attempt_result_
    payload danh cho chinh nguoi thi)."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo Tenant')
        self.bank = QuestionBank.objects.create(tenant=self.tenant, name='Bank demo')
        self.question = Question.objects.create(
            tenant=self.tenant, bank=self.bank, type=Question.Type.SINGLE, stem_html='Câu 1?', points=10,
        )
        self.correct_opt = QuestionOption.objects.create(
            tenant=self.tenant, question=self.question, content_html='Đúng', is_correct=True,
        )
        self.wrong_opt = QuestionOption.objects.create(
            tenant=self.tenant, question=self.question, content_html='Sai', is_correct=False,
        )
        # show_result_mode='score_only'/review_mode='none' - CO Y de chung minh
        # employee_attempts_review KHONG bi gioi han boi 2 cau hinh nay (khac attempt_result_
        # payload).
        self.assessment = Assessment.objects.create(
            tenant=self.tenant, title='Đề 1 câu', status=Assessment.Status.PUBLISHED,
            show_result_mode=Assessment.ShowResultMode.SCORE_ONLY, review_mode=Assessment.ReviewMode.NONE,
        )
        AssessmentQuestion.objects.create(tenant=self.tenant, assessment=self.assessment, question=self.question)
        self.learner_user = User.objects.create_user(
            username='nv1', password='x', tenant=self.tenant, role=User.Role.EMPLOYEE,
        )
        self.employee = Employee.objects.create(tenant=self.tenant, code='NV1', name='NV1', user=self.learner_user)
        AssessmentAssignment.objects.create(tenant=self.tenant, assessment=self.assessment, employee=self.employee)
        self.client = APIClient()

    def _submit(self, option):
        self.client.force_authenticate(self.learner_user)
        attempt_id = self.client.post(reverse('exam-start', args=[self.assessment.id])).data['attempt_id']
        self.client.post(
            reverse('exam-attempt-answer', args=[attempt_id]),
            {'question': self.question.id, 'response': {'option_id': option.id}}, format='json',
        )
        self.client.post(reverse('exam-attempt-submit', args=[attempt_id]))
        return attempt_id

    def test_returns_full_detail_regardless_of_show_result_mode(self):
        from .services import employee_attempts_review

        self._submit(self.correct_opt)
        rows = employee_attempts_review(self.employee)
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row['assessment_title'], 'Đề 1 câu')
        self.assertTrue(row['passed'])
        self.assertEqual(len(row['details']), 1)
        detail = row['details'][0]
        self.assertEqual(detail['stem_html'], 'Câu 1?')
        self.assertTrue(detail['is_correct'])
        self.assertEqual(detail['response'], {'option_id': self.correct_opt.id})
        option_ids = {o['id'] for o in detail['options']}
        self.assertEqual(option_ids, {self.correct_opt.id, self.wrong_opt.id})
        # Options tra ve KEM is_correct - khac attempt_question_payload (danh cho nguoi dang
        # thi, KHONG kem dap an dung) - day la diem khac biet CHINH cua ham nay.
        correct_flags = {o['id']: o['is_correct'] for o in detail['options']}
        self.assertTrue(correct_flags[self.correct_opt.id])
        self.assertFalse(correct_flags[self.wrong_opt.id])

    def test_wrong_answer_marked_incorrect(self):
        from .services import employee_attempts_review

        self._submit(self.wrong_opt)
        row = employee_attempts_review(self.employee)[0]
        self.assertFalse(row['passed'])
        self.assertFalse(row['details'][0]['is_correct'])
        self.assertEqual(row['details'][0]['response'], {'option_id': self.wrong_opt.id})

    def test_in_progress_attempt_excluded(self):
        from .services import employee_attempts_review

        self.client.force_authenticate(self.learner_user)
        self.client.post(reverse('exam-start', args=[self.assessment.id]))
        self.assertEqual(employee_attempts_review(self.employee), [])
