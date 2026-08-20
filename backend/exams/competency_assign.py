"""
Gan nang luc hang loat cho cau hoi qua Excel round-trip (Prompt_GanNangLuc_CauHoi_Excel.md).
Xuat: 1 sheet "Gan nang luc" (Ma cau hoi + noi dung/chu de/dang de doi chieu + cot NANG LUC
trong/dang co de dien) + sheet "DanhMuc_NangLuc" (danh sach ten Competency cua tenant) dung lam
nguon cho DATA VALIDATION dropdown tren cot NANG LUC - giong file mau
MAU_Import_GanNangLuc_CauHoi.xlsx. Nhap: doc lai file da dien, khop CAU HOI theo Ma cau hoi (id),
khop TEN NANG LUC da chuan hoa (tai dung dashboard.services._normalize_competency_name - CUNG 1
cach chuan hoa voi luc import khung nang luc, tranh lech logic).
"""
from .models import Question

SHEET_DATA = 'Gan nang luc'
SHEET_CATALOG = 'DanhMuc_NangLuc'
SHEET_GUIDE = 'Huong dan'
HEADER = ['Mã câu hỏi', 'Nội dung câu hỏi', 'Chủ đề', 'Dạng', 'NĂNG LỰC (chọn)']


def export_workbook(tenant, questions):
    """questions: queryset Question (da loc san theo bo loc cua admin - bank/type/difficulty...).
    Tra ve openpyxl.Workbook (chua save) - view tu quyet dinh xuat ra HttpResponse."""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    from dashboard.models import Competency

    questions = list(questions.select_related('bank', 'competency').order_by('bank__name', 'id'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_DATA
    ws.append(HEADER)
    for q in questions:
        ws.append([
            q.id, q.stem_html, q.bank.name if q.bank_id else '', q.get_type_display(),
            q.competency.name if q.competency_id else '',
        ])

    competencies = list(
        Competency.objects.filter(tenant=tenant).select_related('group').order_by('group__order', 'order')
    )
    ws_cat = wb.create_sheet(SHEET_CATALOG)
    ws_cat.append(['Năng lực'])
    for c in competencies:
        ws_cat.append([c.name])

    last_row = max(len(questions) + 1, 1000)
    catalog_last_row = max(len(competencies) + 1, 2)
    dv = DataValidation(
        type='list', formula1=f"={SHEET_CATALOG}!$A$2:$A${catalog_last_row}", allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f'E2:E{last_row}')

    ws_guide = wb.create_sheet(SHEET_GUIDE)
    for line in [
        'CÁCH DÙNG FILE GÁN NĂNG LỰC',
        '',
        "1. Cột 'Mã câu hỏi'/'Nội dung câu hỏi'/'Chủ đề'/'Dạng' là để đối chiếu — KHÔNG sửa.",
        "2. Điền cột 'NĂNG LỰC (chọn)' — bấm vào ô sẽ có dropdown chọn năng lực (không cần gõ tay).",
        '3. Nhập lại file → hệ thống gán năng lực cho câu hỏi theo Mã câu hỏi.',
        '4. Để trống ô NĂNG LỰC = giữ nguyên năng lực hiện tại (không xoá).',
    ]:
        ws_guide.append([line])

    return wb


def _find_header_row(rows, max_scan=3):
    for i, row in enumerate(rows[:max_scan]):
        cells = [(str(c).strip() if c is not None else '') for c in row]
        if 'Mã câu hỏi' in cells and 'NĂNG LỰC (chọn)' in cells:
            return i, {name: j for j, name in enumerate(cells) if name}
    return None, None


def parse_import_workbook(path_or_file, sheet_name=SHEET_DATA):
    """Doc file da dien -> list[{'excel_row', 'question_id_raw', 'competency_raw'}]. THUAN TUY
    doc file, KHONG dong DB (khop cau hoi/nang luc thuc su lam o apply_competency_assignments)."""
    import openpyxl

    wb = openpyxl.load_workbook(path_or_file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong file (có: {', '.join(wb.sheetnames)}).")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_idx, col = _find_header_row(rows)
    if header_idx is None:
        raise ValueError("Không tìm được dòng tiêu đề (cần cột 'Mã câu hỏi' và 'NĂNG LỰC (chọn)').")

    id_col = col['Mã câu hỏi']
    competency_col = col['NĂNG LỰC (chọn)']

    out = []
    for i, row in enumerate(rows[header_idx + 1:]):
        excel_row = header_idx + 2 + i  # so dong Excel THAT (1-based, tinh ca header) de bao loi de doi chieu
        if row is None or not any(c is not None and str(c).strip() for c in row):
            continue
        id_raw = row[id_col] if id_col < len(row) else None
        competency_raw = row[competency_col] if competency_col < len(row) else None
        out.append({
            'excel_row': excel_row,
            'question_id_raw': '' if id_raw is None else str(id_raw).strip(),
            'competency_raw': '' if competency_raw is None else str(competency_raw).strip(),
        })
    return out


def build_competency_index(tenant):
    """{ten_da_chuan_hoa: [Competency, ...]} - dung chung boi apply_competency_assignments va
    exams/cls_import.py (cot 'Nang luc' tuy chon luc import CLS)."""
    from dashboard.models import Competency
    from dashboard.services import _normalize_competency_name

    index = {}
    for c in Competency.objects.filter(tenant=tenant):
        index.setdefault(_normalize_competency_name(c.name), []).append(c)
    return index


def apply_competency_assignments(tenant, raw_rows, dry_run=True):
    """Khop tung dong voi Question (theo id, trong tenant) + Competency (theo ten da chuan hoa).
    O NANG LUC trong = giu nguyen (khong phai loi). Ten khong khop DUNG 1 Competency (0 hoac >1
    ket qua) = LOI dong do, KHONG tao Competency moi, KHONG doi Question.competency dong do.
    dry_run=True: chi tinh toan + tra ve, khong ghi DB. Idempotent tu nhien (set field, khong
    tao ban ghi) - chay lai voi cung du lieu cho ra cung ket qua."""
    from dashboard.services import _normalize_competency_name

    competency_index = build_competency_index(tenant)
    stats = {'will_assign': 0, 'unchanged_blank': 0, 'errors': 0}
    errors = []
    to_assign = []  # [(question, competency)] - chi dung khi khong dry_run

    for row in raw_rows:
        excel_row = row['excel_row']
        id_raw = row['question_id_raw']
        if not id_raw:
            errors.append({'row': excel_row, 'reason': 'Thiếu Mã câu hỏi'})
            stats['errors'] += 1
            continue
        try:
            question_id = int(id_raw)
        except ValueError:
            errors.append({'row': excel_row, 'reason': f"Mã câu hỏi không hợp lệ: '{id_raw}'"})
            stats['errors'] += 1
            continue

        question = Question.objects.filter(tenant=tenant, pk=question_id).first()
        if not question:
            errors.append({'row': excel_row, 'reason': f'Không tìm thấy câu hỏi id={question_id}'})
            stats['errors'] += 1
            continue

        competency_raw = row['competency_raw']
        if not competency_raw:
            stats['unchanged_blank'] += 1
            continue

        matches = competency_index.get(_normalize_competency_name(competency_raw)) or []
        if len(matches) == 0:
            errors.append({'row': excel_row, 'reason': f"Không khớp năng lực nào tên '{competency_raw}'"})
            stats['errors'] += 1
            continue
        if len(matches) > 1:
            errors.append({
                'row': excel_row,
                'reason': f"Tên năng lực '{competency_raw}' trùng ở nhiều nhóm — không rõ chọn nhóm nào",
            })
            stats['errors'] += 1
            continue

        stats['will_assign'] += 1
        to_assign.append((question, matches[0]))

    if not dry_run:
        for question, competency in to_assign:
            question.competency = competency
            question.save(update_fields=['competency'])

    return {'stats': stats, 'errors': errors}
